"""
Routes for Employee Bulk Upload functionality
Requirement: EMP-BULK-001, EMP-BULK-002, EMP-BULK-003
"""
from flask import render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import current_user
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import os
import time as pytime

from app import app, db
from core.auth import require_login, require_role
from core.models import Employee, User, Role, Department, WorkingHours, WorkSchedule, Company
from core.utils import parse_date, validate_nric, generate_employee_id
from core.constants import DEFAULT_USER_PASSWORD


@app.route('/employees/bulk-upload')
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def employee_bulk_upload():
    """Employee Bulk Upload page with download template and upload options"""
    return render_template('employees/bulk_upload.html')


@app.route('/employees/bulk-upload/download-template')
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def download_employee_template():
    """Download Excel template for bulk employee upload"""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Template"
    
    # Define headers with mandatory field indicators
    headers = [
        ('Employee ID*', 'Auto-generated if left blank'),
        ('First Name*', 'Required'),
        ('Last Name*', 'Required'),
        ('Email*', 'Required - must be unique'),
        ('Phone', 'Optional'),
        ('NRIC/Passport*', 'Required - must be unique (e.g., S1234567D)'),
        ('Date of Birth', 'Format: YYYY-MM-DD (e.g., 1990-01-15)'),
        ('Gender', 'Male/Female'),
        ('Nationality', 'e.g., Singaporean, Malaysian, Indian'),
        ('Address', 'Full address'),
        ('Postal Code', 'e.g., 123456'),
        ('Company Code*', 'Required - Company code from system'),
        ('User Role*', 'Required - Super Admin/Admin/HR Manager/Manager/User'),
        ('Department', 'Department name'),
        ('Manager Employee ID', 'Employee ID of reporting manager'),
        ('Hire Date*', 'Required - Format: YYYY-MM-DD'),
        ('Employment Type*', 'Required - Full-time/Part-time/Contract/Intern'),
        ('Work Permit Type*', 'Required - Citizen/PR/Work Permit/S Pass/Employment Pass'),
        ('Work Permit Expiry', 'Format: YYYY-MM-DD (leave blank for Citizen/PR)'),
        ('Basic Salary*', 'Required - Monthly salary in SGD'),
        ('Allowances', 'Monthly allowances in SGD (default: 0)'),
        ('Hourly Rate', 'Hourly rate for overtime calculation'),
        ('Employee CPF Rate', 'Percentage (default: 20.00)'),
        ('Employer CPF Rate', 'Percentage (default: 17.00)'),
        ('CPF Account', 'CPF account number'),
        ('Bank Name', 'e.g., DBS, OCBC, UOB'),
        ('Bank Account', 'Bank account number'),
        ('Account Holder Name', 'Name as per bank account'),
        ('Working Hours Name', 'Working hours configuration name'),
        ('Work Schedule Name', 'Work schedule name'),
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    mandatory_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, (header, description) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        # Highlight mandatory fields
        if '*' in header:
            cell.fill = mandatory_fill
            cell.font = Font(bold=True, color="000000", size=11)
        else:
            cell.fill = header_fill
        
        # Write description in row 2
        desc_cell = ws.cell(row=2, column=col_num)
        desc_cell.value = description
        desc_cell.font = Font(italic=True, size=9, color="666666")
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        desc_cell.border = border
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40
    
    # Add sample data row
    sample_data = [
        'EMP001',  # Employee ID
        'John',  # First Name
        'Doe',  # Last Name
        'john.doe@example.com',  # Email
        '+65 9123 4567',  # Phone
        'S1234567D',  # NRIC
        '1990-01-15',  # DOB
        'Male',  # Gender
        'Singaporean',  # Nationality
        '123 Main Street, Singapore',  # Address
        '123456',  # Postal Code
        'COMP001',  # Company Code
        'User',  # User Role
        'Engineering',  # Department
        'EMP000',  # Manager Employee ID
        '2024-01-01',  # Hire Date
        'Full-time',  # Employment Type
        'Citizen',  # Work Permit Type
        '',  # Work Permit Expiry
        '5000.00',  # Basic Salary
        '500.00',  # Allowances
        '30.00',  # Hourly Rate
        '20.00',  # Employee CPF Rate
        '17.00',  # Employer CPF Rate
        '',  # CPF Account
        'DBS',  # Bank Name
        '1234567890',  # Bank Account
        'John Doe',  # Account Holder Name
        'Standard Hours',  # Working Hours
        'Day Shift',  # Work Schedule
    ]
    
    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = value
        cell.font = Font(italic=True, color="999999")
        cell.border = border
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Employee Bulk Upload Instructions", ""],
        ["", ""],
        ["Step 1:", "Download this template"],
        ["Step 2:", "Fill in employee data starting from row 3 (row 1 is headers, row 2 is descriptions)"],
        ["Step 3:", "Fields marked with * are mandatory"],
        ["Step 4:", "Mandatory fields are highlighted in YELLOW"],
        ["Step 5:", "Follow the date format: YYYY-MM-DD (e.g., 2024-01-15)"],
        ["Step 6:", "Employee ID will be auto-generated if left blank"],
        ["Step 7:", "Email and NRIC must be unique across all employees"],
        ["Step 8:", "Save the file and upload it using the 'Upload Excel' button"],
        ["", ""],
        ["Important Notes:", ""],
        ["", "• Company Code must exist in the system"],
        ["", "• User Role must be one of: Super Admin, Admin, HR Manager, Manager, User"],
        ["", "• Employment Type: Full-time, Part-time, Contract, or Intern"],
        ["", "• Work Permit Type: Citizen, PR, Work Permit, S Pass, or Employment Pass"],
        ["", "• All monetary values should be in SGD"],
        ["", "• CPF rates are in percentage (e.g., 20.00 for 20%)"],
        ["", "• Manager Employee ID should be an existing employee's ID"],
        ["", ""],
        ["For any issues, contact your system administrator.", ""],
    ]
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1).value = col1
        ws_instructions.cell(row=row_num, column=2).value = col2
        if row_num == 1:
            ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="4472C4")
        elif "Step" in col1 or "Important" in col1:
            ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    
    ws_instructions.column_dimensions['A'].width = 30
    ws_instructions.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"Employee_Bulk_Upload_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/employees/bulk-upload/upload', methods=['POST'])
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def upload_employee_excel():
    """Process uploaded Excel file and create/update employees"""
    # Import locally to avoid circular import issues if any
    from core.utils import get_company_employee_id
    
    if 'excel_file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('employee_bulk_upload'))
    
    file = request.files['excel_file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('employee_bulk_upload'))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Invalid file format. Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('employee_bulk_upload'))
    
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Skip header rows (row 1 and 2)
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        
        if not rows:
            flash('No data found in the Excel file', 'error')
            return redirect(url_for('employee_bulk_upload'))
        
        success_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=3):
            # Skip empty rows
            if not any(row):
                continue
            
            try:
                # Extract basic data first
                provided_employee_id = str(row[0]).strip() if row[0] else None
                first_name = row[1]
                last_name = row[2]
                email = row[3]
                phone = row[4] if len(row) > 4 else None
                nric = row[5] if len(row) > 5 else None
                date_of_birth = row[6] if len(row) > 6 else None
                gender = row[7] if len(row) > 7 else None
                nationality = row[8] if len(row) > 8 else None
                address = row[9] if len(row) > 9 else None
                postal_code = row[10] if len(row) > 10 else None
                company_code = row[11] if len(row) > 11 else None
                # ... check indexes for other fields ...
                user_role_name = row[12] if len(row) > 12 else 'User'
                department = str(row[13]).strip() if len(row) > 13 and row[13] else None
                manager_employee_id = row[14] if len(row) > 14 else None
                hire_date = row[15] if len(row) > 15 else None
                employment_type = row[16] if len(row) > 16 else None
                work_permit_type = row[17] if len(row) > 17 else None
                work_permit_expiry = row[18] if len(row) > 18 else None
                basic_salary = row[19] if len(row) > 19 else None
                allowances = row[20] if len(row) > 20 else 0
                hourly_rate = row[21] if len(row) > 21 else None
                employee_cpf_rate = row[22] if len(row) > 22 else 20.00
                employer_cpf_rate = row[23] if len(row) > 23 else 17.00
                cpf_account = row[24] if len(row) > 24 else None
                bank_name = row[25] if len(row) > 25 else None
                bank_account = row[26] if len(row) > 26 else None
                account_holder_name = row[27] if len(row) > 27 else None
                working_hours_name = row[28] if len(row) > 28 else None
                work_schedule_name = row[29] if len(row) > 29 else None
                
                # Validate mandatory fields
                if not all([first_name, last_name, email, nric, company_code, hire_date, employment_type, work_permit_type, basic_salary]):
                    errors.append(f"Row {row_num}: Missing mandatory fields")
                    error_count += 1
                    continue

                # Get company FIRST (needed for ID generation)
                company = Company.query.filter_by(code=company_code).first()
                if not company:
                    errors.append(f"Row {row_num}: Company code {company_code} not found")
                    error_count += 1
                    continue
                
                # NOW handle Employee ID Logic
                is_update = False
                employee = None
                
                if provided_employee_id:
                    # User provided an ID - check if it exists
                    employee = Employee.query.filter_by(employee_id=provided_employee_id).first()
                    if employee:
                        is_update = True
                        employee_id = provided_employee_id
                    else:
                        # Provided ID but doesn't exist - use it for new record
                        employee_id = provided_employee_id
                else:
                    # No ID provided - GENERATE using company sequence
                    employee_id = get_company_employee_id(company.id, company.code, db.session)
                
                # Check for duplicate Email (Update Logic)
                existing_email = Employee.query.filter_by(email=email).first()
                if existing_email:
                    if not is_update or (is_update and existing_email.id != employee.id):
                        errors.append(f"Row {row_num}: Email {email} already exists for another employee")
                        error_count += 1
                        continue
                
                # Check for duplicate NRIC (Update Logic)
                if nric:
                    existing_nric = Employee.query.filter_by(nric=nric).first()
                    if existing_nric:
                        if not is_update or (is_update and existing_nric.id != employee.id):
                            errors.append(f"Row {row_num}: NRIC {nric} already exists for another employee")
                            error_count += 1
                            continue
                
                # Get user role
                user_role = Role.query.filter_by(name=user_role_name).first()
                if not user_role:
                    errors.append(f"Row {row_num}: User role {user_role_name} not found")
                    error_count += 1
                    continue
                
                # Get manager if specified
                manager_id = None
                if manager_employee_id:
                    manager = Employee.query.filter_by(employee_id=manager_employee_id).first()
                    if manager:
                        manager_id = manager.id
                    else:
                        errors.append(f"Row {row_num}: Warning - Manager {manager_employee_id} not found, skipping manager assignment")
                
                # Get working hours
                working_hours_id = None
                if working_hours_name:
                    working_hours = WorkingHours.query.filter_by(name=working_hours_name).first()
                    if working_hours:
                        working_hours_id = working_hours.id
                
                # Get work schedule
                work_schedule_id = None
                if work_schedule_name:
                    work_schedule = WorkSchedule.query.filter_by(name=work_schedule_name).first()
                    if work_schedule:
                        work_schedule_id = work_schedule.id
                
                # Parse dates
                if isinstance(date_of_birth, str):
                    date_of_birth = parse_date(date_of_birth)
                if isinstance(hire_date, str):
                    hire_date = parse_date(hire_date)
                if isinstance(work_permit_expiry, str) and work_permit_expiry:
                    work_permit_expiry = parse_date(work_permit_expiry)
                
                if is_update:
                    # Update existing employee
                    employee.first_name = first_name
                    employee.last_name = last_name
                    employee.email = email
                    employee.phone = phone
                    employee.nric = nric
                    employee.date_of_birth = date_of_birth
                    employee.gender = gender
                    employee.nationality = nationality
                    employee.address = address
                    employee.postal_code = postal_code
                    employee.company_id = company.id
                    employee.department = department
                    employee.manager_id = manager_id
                    employee.hire_date = hire_date
                    employee.employment_type = employment_type
                    employee.work_permit_type = work_permit_type
                    employee.work_permit_expiry = work_permit_expiry
                    employee.basic_salary = float(basic_salary)
                    employee.allowances = float(allowances) if allowances else 0
                    employee.hourly_rate = float(hourly_rate) if hourly_rate else None
                    employee.employee_cpf_rate = float(employee_cpf_rate) if employee_cpf_rate else 20.00
                    employee.employer_cpf_rate = float(employer_cpf_rate) if employer_cpf_rate else 17.00
                    employee.cpf_account = cpf_account
                    employee.bank_name = bank_name
                    employee.bank_account = bank_account
                    employee.account_holder_name = account_holder_name
                    employee.working_hours_id = working_hours_id
                    employee.work_schedule_id = work_schedule_id
                    
                    # Update linked User
                    if employee.user:
                        employee.user.email = email
                        employee.user.first_name = first_name
                        employee.user.last_name = last_name
                        employee.user.role_id = user_role.id
                    
                    updated_count += 1
                else:
                    # Create new employee
                    employee = Employee(
                        employee_id=employee_id,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone=phone,
                        nric=nric,
                        date_of_birth=date_of_birth,
                        gender=gender,
                        nationality=nationality,
                        address=address,
                        postal_code=postal_code,
                        company_id=company.id,
                        department=department,
                        manager_id=manager_id,
                        hire_date=hire_date,
                        employment_type=employment_type,
                        work_permit_type=work_permit_type,
                        work_permit_expiry=work_permit_expiry,
                        basic_salary=float(basic_salary),
                        allowances=float(allowances) if allowances else 0,
                        hourly_rate=float(hourly_rate) if hourly_rate else None,
                        employee_cpf_rate=float(employee_cpf_rate) if employee_cpf_rate else 20.00,
                        employer_cpf_rate=float(employer_cpf_rate) if employer_cpf_rate else 17.00,
                        cpf_account=cpf_account,
                        bank_name=bank_name,
                        bank_account=bank_account,
                        account_holder_name=account_holder_name,
                        working_hours_id=working_hours_id,
                        work_schedule_id=work_schedule_id,
                        organization_id=current_user.organization_id,
                        is_active=True,
                        created_by=current_user.username
                    )
                    
                    db.session.add(employee)
                    db.session.flush()
                    
                    # Create user account
                    username = employee.employee_id
                    if User.query.filter_by(username=username).first():
                        # Should not happen if ID sequence is correct, but safe fallback
                        if 'EMP' in username and len(username) > 10: 
                             # If it was a timestamp fallback, retry with new timestamp? 
                             # Unlikely to hit this if we use get_company_employee_id
                             pass
                        raise ValueError(f"User account '{username}' already exists")
                    
                    user = User(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        organization_id=current_user.organization_id,
                        role_id=user_role.id,
                        is_active=True
                    )
                    user.set_password(DEFAULT_USER_PASSWORD)
                    user.must_reset_password = True
                    
                    db.session.add(user)
                    db.session.flush()
                    
                    employee.user_id = user.id
                    success_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
                db.session.rollback()
                continue
        
        # Commit
        if success_count > 0 or updated_count > 0:
            db.session.commit()
            msg = []
            if success_count > 0:
                msg.append(f'Created {success_count} new employee(s)')
            if updated_count > 0:
                msg.append(f'Updated {updated_count} existing employee(s)')
            flash(' & '.join(msg), 'success')
        
        if error_count > 0:
            flash(f'{error_count} row(s) failed. See errors below.', 'warning')
            for error in errors[:10]:
                flash(error, 'error')
            if len(errors) > 10:
                flash(f'... and {len(errors) - 10} more errors', 'error')
        
        return redirect(url_for('employee_list'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing Excel file: {str(e)}', 'error')
        return redirect(url_for('employee_bulk_upload'))


@app.route('/employees/export')
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def export_employees():
    """Export employees to Excel in bulk upload format"""
    # Import Tenant to support search filtering by tenant name matching employee_list logic
    from core.models import Tenant

    search = request.args.get('search', '', type=str)
    status = request.args.get('status', 'active', type=str).lower()
    department = request.args.get('department', '', type=str)
    designation_id = request.args.get('designation_id', '', type=str)
    employment_type = request.args.get('employment_type', '', type=str)
    company_id = request.args.get('company_id', '', type=str)

    # Base Query
    query = db.session.query(Employee).join(
        Company, Employee.company_id == Company.id
    ).outerjoin(
        Tenant, Company.tenant_id == Tenant.id
    )

    # 1. Status Filter
    if status == 'active':
        query = query.filter(Employee.is_active == True)
    elif status == 'inactive':
        query = query.filter(Employee.is_active == False)
    
    # 2. Search Filter
    if search:
        query = query.filter(
            db.or_(Employee.first_name.ilike(f'%{search}%'),
                   Employee.last_name.ilike(f'%{search}%'),
                   Employee.employee_id.ilike(f'%{search}%'),
                   Employee.email.ilike(f'%{search}%'),
                   Company.name.ilike(f'%{search}%'),
                   Tenant.name.ilike(f'%{search}%')))

    # 3. Specific Field Filters
    if department:
        query = query.filter(Employee.department == department)
    
    if designation_id and designation_id.isdigit():
        query = query.filter(Employee.designation_id == int(designation_id))

    if employment_type:
        query = query.filter(Employee.employment_type == employment_type)

    if company_id:
        try:
            query = query.filter(Employee.company_id == company_id)
        except Exception:
            pass

    # 4. Role-based Access Control
    user_role = current_user.role.name if current_user.role else None
    if user_role in ['HR Manager', 'Tenant Admin']:
        accessible_companies = current_user.get_accessible_companies()
        company_ids = [c.id for c in accessible_companies]
        query = query.filter(Employee.company_id.in_(company_ids))

    # Order by Employee ID
    employees = query.order_by(Employee.employee_id).all()

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Export"

    # Reuse headers from download_employee_template logic for consistency
    headers = [
        ('Employee ID*', 'Auto-generated if left blank'),
        ('First Name*', 'Required'),
        ('Last Name*', 'Required'),
        ('Email*', 'Required - must be unique'),
        ('Phone', 'Optional'),
        ('NRIC/Passport*', 'Required - must be unique (e.g., S1234567D)'),
        ('Date of Birth', 'Format: YYYY-MM-DD (e.g., 1990-01-15)'),
        ('Gender', 'Male/Female'),
        ('Nationality', 'e.g., Singaporean, Malaysian, Indian'),
        ('Address', 'Full address'),
        ('Postal Code', 'e.g., 123456'),
        ('Company Code*', 'Required - Company code from system'),
        ('User Role*', 'Required - Super Admin/Admin/HR Manager/Manager/User'),
        ('Department', 'Department name'),
        ('Manager Employee ID', 'Employee ID of reporting manager'),
        ('Hire Date*', 'Required - Format: YYYY-MM-DD'),
        ('Employment Type*', 'Required - Full-time/Part-time/Contract/Intern'),
        ('Work Permit Type*', 'Required - Citizen/PR/Work Permit/S Pass/Employment Pass'),
        ('Work Permit Expiry', 'Format: YYYY-MM-DD (leave blank for Citizen/PR)'),
        ('Basic Salary*', 'Required - Monthly salary in SGD'),
        ('Allowances', 'Monthly allowances in SGD (default: 0)'),
        ('Hourly Rate', 'Hourly rate for overtime calculation'),
        ('Employee CPF Rate', 'Percentage (default: 20.00)'),
        ('Employer CPF Rate', 'Percentage (default: 17.00)'),
        ('CPF Account', 'CPF account number'),
        ('Bank Name', 'e.g., DBS, OCBC, UOB'),
        ('Bank Account', 'Bank account number'),
        ('Account Holder Name', 'Name as per bank account'),
        ('Working Hours Name', 'Working hours configuration name'),
        ('Work Schedule Name', 'Work schedule name'),
    ]

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    mandatory_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write Headers matches template
    for col_num, (header, description) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        if '*' in header:
            cell.fill = mandatory_fill
            cell.font = Font(bold=True, color="000000", size=11)
        else:
            cell.fill = header_fill

        desc_cell = ws.cell(row=2, column=col_num)
        desc_cell.value = description
        desc_cell.font = Font(italic=True, size=9, color="666666")
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        desc_cell.border = border
        
        ws.column_dimensions[get_column_letter(col_num)].width = 20
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40

    # Write Data
    for row_num, emp in enumerate(employees, 3):
        # Resolve relations
        manager_emp_id = emp.manager.employee_id if emp.manager else ''
        company_code = emp.company.code if emp.company else ''
        role_name = emp.user.role.name if emp.user and emp.user.role else ''
        working_hours_name = emp.working_hours.name if emp.working_hours else ''
        work_schedule_name = emp.work_schedule.name if emp.work_schedule else ''

        row_data = [
            emp.employee_id,
            emp.first_name,
            emp.last_name,
            emp.email,
            emp.phone,
            emp.nric,
            emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else '',
            emp.gender,
            emp.nationality,
            emp.address,
            emp.postal_code,
            company_code,
            role_name,
            emp.department,
            manager_emp_id,
            emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '',
            emp.employment_type,
            emp.work_permit_type,
            emp.work_permit_expiry.strftime('%Y-%m-%d') if emp.work_permit_expiry else '',
            str(emp.basic_salary) if emp.basic_salary is not None else '',
            str(emp.allowances) if emp.allowances is not None else '',
            str(emp.hourly_rate) if emp.hourly_rate is not None else '',
            str(emp.employee_cpf_rate) if emp.employee_cpf_rate is not None else '',
            str(emp.employer_cpf_rate) if emp.employer_cpf_rate is not None else '',
            emp.cpf_account,
            emp.bank_name,
            emp.bank_account,
            emp.account_holder_name,
            working_hours_name,
            work_schedule_name
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

    # Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Employees_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
