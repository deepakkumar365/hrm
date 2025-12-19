"""
Routes for Access Control Management
Allows Super Admin to manage role-based access to modules, menus, and sub-menus
"""

from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for
from flask_login import current_user, login_required
from sqlalchemy import and_, or_
from datetime import datetime
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app import app, db
from core.auth import require_role
from core.models import (
    RoleAccessControl, UserRoleMapping, User, Role, 
    Organization, Company, AuditLog, UserCompanyAccess
)
from core.utils import check_permission

# Create blueprint
access_control_bp = Blueprint('access_control', __name__, url_prefix='/access-control')

# =====================================================================
# DEFAULT MODULES, MENUS & SUB-MENUS DATA
# =====================================================================

DEFAULT_MODULES = {
    'Payroll': {
        'Payroll Management': ['Payroll List', 'Payroll Generation', 'Payroll Approval', 'Payroll History'],
        'Payslip Management': ['View Payslips', 'Download Payslips', 'Print Payslips'],
        'Payroll Reports': ['Salary Reports', 'Tax Reports', 'Deduction Reports'],
    },
    'Attendance': {
        'Attendance Management': ['Mark Attendance', 'Attendance List', 'Attendance Reports', 'Bulk Upload'],
        'Leave Management': ['Apply Leave', 'Leave Approval', 'Leave Balance', 'Leave Reports'],
    },
    'Employees': {
        'Employee Management': ['View Employees', 'Add Employee', 'Edit Employee', 'Employee List'],
        'Employee Documents': ['Upload Documents', 'View Documents', 'Download Documents'],
        'Employee Reports': ['Employee Directory', 'Employee Summary'],
    },
    'Claims': {
        'Expense Claims': ['Submit Claim', 'Claim Approval', 'Claim History'],
        'Claim Reports': ['Claim Summary', 'Claim Analysis'],
    },
    'Appraisals': {
        'Appraisal Management': ['Create Appraisal', 'View Appraisals', 'Submit Appraisal'],
        'Appraisal Reports': ['Appraisal Summary', 'Performance Reports'],
    },
    'Admin Settings': {
        'Access Control Configuration': ['View Access Matrix', 'Edit Access Matrix', 'Export Matrix', 'Import Matrix'],
        'User Role Mapping': ['Map Roles', 'Manage User Roles', 'Manage Company Access'],
        'Master Data': ['Manage Roles', 'Manage Departments', 'Manage Designations'],
    },
}

ACCESS_LEVEL_OPTIONS = ['Editable', 'View Only', 'Hidden']
ROLE_NAMES = ['Super Admin', 'Tenant Admin', 'HR Manager', 'Employee']


def log_audit(action, resource_type, resource_id, changes, status='Success'):
    """Log access control changes to audit log"""
    try:
        audit_entry = AuditLog(
            user_id=current_user.id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            changes=changes,
            status=status,
            created_at=datetime.now()
        )
        db.session.add(audit_entry)
        db.session.commit()
    except Exception as e:
        print(f"Audit log error: {str(e)}")


def initialize_access_control_matrix():
    """Initialize default access control matrix if not exists"""
    try:
        # Check if any access control records exist
        if RoleAccessControl.query.count() > 0:
            return
        
        # Create default entries
        for module_name, menus in DEFAULT_MODULES.items():
            for menu_name, sub_menus in menus.items():
                for sub_menu_name in sub_menus:
                    # Set default access levels
                    if module_name == 'Admin Settings':
                        # Admin Settings: Super Admin has full access
                        # Tenant Admin can access Master Data (Roles, Departments, Designations)
                        # HR Manager and Employee cannot access admin settings
                        if menu_name == 'Master Data':
                            # Tenant Admin can manage master data
                            access_level = RoleAccessControl(
                                module_name=module_name,
                                menu_name=menu_name,
                                sub_menu_name=sub_menu_name,
                                super_admin_access='Editable',
                                tenant_admin_access='Editable',
                                hr_manager_access='Hidden',
                                employee_access='Hidden',
                                created_by='system'
                            )
                        else:
                            # Other admin settings only for Super Admin
                            access_level = RoleAccessControl(
                                module_name=module_name,
                                menu_name=menu_name,
                                sub_menu_name=sub_menu_name,
                                super_admin_access='Editable',
                                tenant_admin_access='Hidden',
                                hr_manager_access='Hidden',
                                employee_access='Hidden',
                                created_by='system'
                            )
                    elif module_name in ['Payroll', 'Appraisals']:
                        # Limited access
                        access_level = RoleAccessControl(
                            module_name=module_name,
                            menu_name=menu_name,
                            sub_menu_name=sub_menu_name,
                            super_admin_access='Editable',
                            tenant_admin_access='Editable',
                            hr_manager_access='View Only',
                            employee_access='View Only',
                            created_by='system'
                        )
                    else:
                        # Default broader access
                        access_level = RoleAccessControl(
                            module_name=module_name,
                            menu_name=menu_name,
                            sub_menu_name=sub_menu_name,
                            super_admin_access='Editable',
                            tenant_admin_access='Editable',
                            hr_manager_access='View Only',
                            employee_access='View Only',
                            created_by='system'
                        )
                    db.session.add(access_level)
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error initializing access control matrix: {str(e)}")


# =====================================================================
# ROUTE: View Access Control Matrix
# =====================================================================

@access_control_bp.route('/matrix', methods=['GET'])
@login_required
@require_role('Super Admin')
def view_access_matrix():
    """Display role-based access control matrix"""
    try:
        # Initialize matrix on first load
        initialize_access_control_matrix()
        
        # Fetch all access control records
        access_controls = RoleAccessControl.query.order_by(
            RoleAccessControl.module_name,
            RoleAccessControl.menu_name,
            RoleAccessControl.sub_menu_name
        ).all()
        
        # Organize data by module and menu
        matrix_data = {}
        for ac in access_controls:
            if ac.module_name not in matrix_data:
                matrix_data[ac.module_name] = {}
            if ac.menu_name not in matrix_data[ac.module_name]:
                matrix_data[ac.module_name][ac.menu_name] = []
            matrix_data[ac.module_name][ac.menu_name].append(ac)
        
        return render_template(
            'access_control/access_matrix.html',
            matrix_data=matrix_data,
            access_levels=ACCESS_LEVEL_OPTIONS,
            roles=ROLE_NAMES[:-1]  # Exclude 'Employee' from role columns
        )
    except Exception as e:
        flash(f'Error loading access matrix: {str(e)}', 'danger')
        return redirect(url_for('index'))


# =====================================================================
# ROUTE: Update Access Control Matrix
# =====================================================================

@access_control_bp.route('/matrix/update', methods=['POST'])
@login_required
@require_role('Super Admin')
def update_access_matrix():
    """Update role access levels in the matrix"""
    try:
        data = request.get_json()
        access_id = data.get('access_id')
        role = data.get('role')  # Column name: super_admin_access, tenant_admin_access, etc.
        access_level = data.get('access_level')
        
        if not access_id or not role or access_level not in ACCESS_LEVEL_OPTIONS:
            return jsonify({'success': False, 'message': 'Invalid parameters'}), 400
        
        # Fetch the access control record
        ac = RoleAccessControl.query.get(access_id)
        if not ac:
            return jsonify({'success': False, 'message': 'Access control record not found'}), 404
        
        # Store old value for audit
        old_value = getattr(ac, f'{role}_access')
        
        # Update the access level
        setattr(ac, f'{role}_access', access_level)
        ac.updated_by = current_user.username
        ac.updated_at = datetime.now()
        
        db.session.commit()
        
        # Log the change
        log_audit(
            action='UPDATE_ACCESS_CONTROL',
            resource_type='RoleAccessControl',
            resource_id=access_id,
            changes=json.dumps({
                'field': f'{role}_access',
                'old_value': old_value,
                'new_value': access_level
            })
        )
        
        return jsonify({'success': True, 'message': 'Access level updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# ROUTE: Reset Access Control Matrix
# =====================================================================

@access_control_bp.route('/matrix/reset', methods=['POST'])
@login_required
@require_role('Super Admin')
def reset_access_matrix():
    """Reset access matrix to default values"""
    try:
        # Delete all existing records
        RoleAccessControl.query.delete()
        db.session.commit()
        
        # Re-initialize with defaults
        initialize_access_control_matrix()
        
        # Log the action
        log_audit(
            action='RESET_ACCESS_MATRIX',
            resource_type='RoleAccessControl',
            resource_id='ALL',
            changes='Matrix reset to default values'
        )
        
        return jsonify({'success': True, 'message': 'Access matrix reset to defaults'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# ROUTE: Export Access Matrix as Excel
# =====================================================================

@access_control_bp.route('/matrix/export', methods=['GET'])
@login_required
@require_role('Super Admin')
def export_access_matrix():
    """Export access matrix as Excel file"""
    try:
        access_controls = RoleAccessControl.query.order_by(
            RoleAccessControl.module_name,
            RoleAccessControl.menu_name
        ).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Access Matrix"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add headers
        headers = ['Module', 'Menu', 'Sub-Menu', 'Super Admin', 'Tenant Admin', 'HR Manager', 'Employee']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Add data
        for row, ac in enumerate(access_controls, 2):
            ws.cell(row=row, column=1, value=ac.module_name).border = border
            ws.cell(row=row, column=2, value=ac.menu_name).border = border
            ws.cell(row=row, column=3, value=ac.sub_menu_name or '').border = border
            ws.cell(row=row, column=4, value=ac.super_admin_access).border = border
            ws.cell(row=row, column=5, value=ac.tenant_admin_access).border = border
            ws.cell(row=row, column=6, value=ac.hr_manager_access).border = border
            ws.cell(row=row, column=7, value=ac.employee_access).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log the action
        log_audit(
            action='EXPORT_ACCESS_MATRIX',
            resource_type='RoleAccessControl',
            resource_id='ALL',
            changes='Matrix exported to Excel'
        )
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'access_matrix_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        flash(f'Error exporting matrix: {str(e)}', 'danger')
        return redirect(url_for('access_control.view_access_matrix'))


# =====================================================================
# ROUTE: Import Access Matrix from Excel
# =====================================================================

@access_control_bp.route('/matrix/import', methods=['POST'])
@login_required
@require_role('Super Admin')
def import_access_matrix():
    """Import and update access matrix from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        changes_count = 0
        errors = []
        
        # Process rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                module, menu, sub_menu, super_admin, tenant_admin, hr_manager, employee = row[:7]
                
                # Validate access levels
                for access_level in [super_admin, tenant_admin, hr_manager, employee]:
                    if access_level and access_level not in ACCESS_LEVEL_OPTIONS:
                        raise ValueError(f'Invalid access level: {access_level}')
                
                # Find or create access control record
                ac = RoleAccessControl.query.filter_by(
                    module_name=module,
                    menu_name=menu,
                    sub_menu_name=sub_menu
                ).first()
                
                if ac:
                    # Update existing
                    if super_admin:
                        ac.super_admin_access = super_admin
                    if tenant_admin:
                        ac.tenant_admin_access = tenant_admin
                    if hr_manager:
                        ac.hr_manager_access = hr_manager
                    if employee:
                        ac.employee_access = employee
                    ac.updated_by = current_user.username
                    ac.updated_at = datetime.now()
                    changes_count += 1
                else:
                    # Create new if specified in import
                    if all([module, menu]):
                        ac = RoleAccessControl(
                            module_name=module,
                            menu_name=menu,
                            sub_menu_name=sub_menu,
                            super_admin_access=super_admin or 'Hidden',
                            tenant_admin_access=tenant_admin or 'Hidden',
                            hr_manager_access=hr_manager or 'Hidden',
                            employee_access=employee or 'Hidden',
                            created_by=current_user.username
                        )
                        db.session.add(ac)
                        changes_count += 1
            except Exception as row_error:
                errors.append(f'Row {row_idx}: {str(row_error)}')
        
        db.session.commit()
        
        # Log the action
        log_audit(
            action='IMPORT_ACCESS_MATRIX',
            resource_type='RoleAccessControl',
            resource_id='ALL',
            changes=json.dumps({'changes_count': changes_count, 'errors': errors})
        )
        
        message = f'Successfully imported {changes_count} records'
        if errors:
            message += f'. Errors: {", ".join(errors[:5])}'
        
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# ROUTE: User Role Mapping Management
# =====================================================================

@access_control_bp.route('/user-roles', methods=['GET'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def manage_user_roles():
    """Display user role mapping interface"""
    try:
        # Get users based on role
        if current_user.role and current_user.role.name == 'Super Admin':
            # Super Admin sees all users
            users = User.query.order_by(User.first_name, User.last_name).all()
            companies = Company.query.filter_by(is_active=True).all()
        else:
            # Tenant Admin/HR Manager see only users from their tenant
            current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
            if current_tenant_id:
                # Get all users in the same tenant
                users = db.session.query(User).join(
                    Organization, User.organization_id == Organization.id
                ).filter(
                    Organization.tenant_id == current_tenant_id
                ).order_by(User.first_name, User.last_name).all()
                
                # Get only companies in the same tenant
                companies = Company.query.filter_by(
                    tenant_id=current_tenant_id,
                    is_active=True
                ).all()
            else:
                users = []
                companies = []
        
        roles = Role.query.all()
        
        # Get all user role mappings
        user_mappings = {}
        for user in users:
            user_mappings[user.id] = UserRoleMapping.query.filter_by(
                user_id=user.id,
                is_active=True
            ).all()
        
        return render_template(
            'access_control/user_role_mapping.html',
            users=users,
            roles=roles,
            companies=companies,
            user_mappings=user_mappings
        )
    except Exception as e:
        flash(f'Error loading user role mapping: {str(e)}', 'danger')
        return redirect(url_for('index'))


# =====================================================================
# ROUTE: Add/Update User Role Mapping
# =====================================================================

@access_control_bp.route('/user-roles/save', methods=['POST'])
@login_required
@require_role('Super Admin')
def save_user_role_mapping():
    """Save user role and company mapping"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        role_ids = data.get('role_ids', [])
        company_ids = data.get('company_ids', [])
        
        if not user_id:
            return jsonify({'success': False, 'message': 'User ID is required'}), 400
        
        # Verify user exists
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Delete existing mappings
        UserRoleMapping.query.filter_by(user_id=user_id).delete()
        
        # Create new mappings
        for role_id in role_ids:
            # One mapping per role (optionally with company)
            for company_id in company_ids if company_ids else [None]:
                mapping = UserRoleMapping(
                    user_id=user_id,
                    role_id=role_id,
                    company_id=company_id,
                    is_active=True,
                    created_by=current_user.username
                )
                db.session.add(mapping)
        
        db.session.commit()
        
        # Log the action
        log_audit(
            action='UPDATE_USER_ROLE_MAPPING',
            resource_type='UserRoleMapping',
            resource_id=user_id,
            changes=json.dumps({
                'role_ids': role_ids,
                'company_ids': company_ids
            })
        )
        
        return jsonify({'success': True, 'message': 'User role mapping saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# UTILITY: Check Role Access to Module/Menu
# =====================================================================

def check_module_access(user_role, module_name, menu_name=None, sub_menu_name=None):
    """
    Check if a role has access to a module/menu/sub-menu
    Returns: 'Editable', 'View Only', 'Hidden'
    """
    try:
        query = RoleAccessControl.query.filter_by(module_name=module_name)
        
        if menu_name:
            query = query.filter_by(menu_name=menu_name)
        if sub_menu_name:
            query = query.filter_by(sub_menu_name=sub_menu_name)
        
        ac = query.first()
        if not ac:
            return 'Hidden'  # Default to hidden if not found
        
        # Map role to column name
        role_column_map = {
            'Super Admin': 'super_admin_access',
            'Tenant Admin': 'tenant_admin_access',
            'HR Manager': 'hr_manager_access',
            'Employee': 'employee_access',
        }
        
        column = role_column_map.get(user_role, 'employee_access')
        access_level = getattr(ac, column, 'Hidden')
        
        return access_level
    except Exception as e:
        print(f"Error checking module access: {str(e)}")
        return 'Hidden'


# =====================================================================
# API: Get User Role Mappings
# =====================================================================

@app.route('/api/user-role-mappings/<int:user_id>', methods=['GET'])
@login_required
@require_role('Super Admin')
def get_user_role_mappings(user_id):
    """Get all role mappings for a specific user"""
    try:
        mappings = UserRoleMapping.query.filter_by(
            user_id=user_id,
            is_active=True
        ).all()
        
        result = []
        for mapping in mappings:
            result.append({
                'id': mapping.id,
                'user_id': mapping.user_id,
                'role_id': mapping.role_id,
                'role_name': mapping.role.name if mapping.role else 'Unknown',
                'company_id': str(mapping.company_id) if mapping.company_id else None,
                'company_name': mapping.company.name if mapping.company else None,
            })
        
        return jsonify({'success': True, 'mappings': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# UTILITY: Enforce Access Control in Routes
# =====================================================================

def check_ui_access(user_role, module_name, menu_name=None):
    """
    Check if UI element should be visible based on role access
    Returns True if 'Editable' or 'View Only', False if 'Hidden'
    """
    access_level = check_module_access(user_role, module_name, menu_name)
    return access_level != 'Hidden'


def check_edit_permission(user_role, module_name, menu_name=None):
    """
    Check if user has edit permission for a module
    Returns True only if access level is 'Editable'
    """
    access_level = check_module_access(user_role, module_name, menu_name)
    return access_level == 'Editable'


# =====================================================================
# ROUTES: Manage User Company Access (Multi-Company Feature)
# =====================================================================

@access_control_bp.route('/manage-user-companies', methods=['GET'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def manage_user_companies():
    """Display page to manage which companies users have access to"""
    try:
        # Always filter by tenant - consistency across all roles
        # Get current user's tenant
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        
        if not current_tenant_id:
            flash('Your organization is not assigned to a tenant', 'danger')
            return redirect(url_for('index'))
        
        # Get users from same tenant (excluding self)
        users = User.query.join(
            Organization, User.organization_id == Organization.id
        ).filter(
            Organization.tenant_id == current_tenant_id,
            User.id != current_user.id  # Exclude self
        ).order_by(User.username).all()
        
        # Get companies from same tenant
        companies = Company.query.filter_by(
            tenant_id=current_tenant_id
        ).order_by(Company.name).all()
        
        return render_template(
            'access_control/manage_user_companies.html',
            users=users,
            companies=companies,
            total_users=len(users),
            total_companies=len(companies)
        )
    except Exception as e:
        flash(f'Error loading user company management: {str(e)}', 'danger')
        return redirect(url_for('index'))


@access_control_bp.route('/api/user-companies/<int:user_id>', methods=['GET'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def get_user_companies(user_id):
    """API: Get all companies for a specific user"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Verify user belongs to same tenant as current user (consistent for all roles)
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        user_tenant_id = user.organization.tenant_id if user.organization else None
        
        if current_tenant_id != user_tenant_id:
            return jsonify({'success': False, 'message': 'Unauthorized: User is not in your tenant'}), 403
        
        companies = [
            {
                'id': str(access.company.id),
                'name': access.company.name,
                'access_id': str(access.id),
                'added_at': access.created_at.strftime('%Y-%m-%d %H:%M') if access.created_at else 'N/A'
            }
            for access in user.company_access
        ]
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'role': user.role.name if user.role else 'N/A'
            },
            'companies': companies,
            'total_companies': len(companies)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@access_control_bp.route('/api/add-company-to-user', methods=['POST'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def add_company_to_user():
    """API: Add company access to a user"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        company_id = data.get('company_id')
        
        if not user_id or not company_id:
            return jsonify({'success': False, 'message': 'User ID and Company ID are required'}), 400
        
        user = User.query.get(user_id)
        company = Company.query.get(company_id)
        
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        if not company:
            return jsonify({'success': False, 'message': 'Company not found'}), 404
        
        # Verify tenant access - consistent for all roles
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        user_tenant_id = user.organization.tenant_id if user.organization else None
        company_tenant_id = company.tenant_id
        
        if current_tenant_id != user_tenant_id or current_tenant_id != company_tenant_id:
            return jsonify({'success': False, 'message': 'Unauthorized: User and Company must be in your tenant'}), 403
        
        # Check if user already has access to this company
        existing_access = UserCompanyAccess.query.filter_by(
            user_id=user_id,
            company_id=company_id
        ).first()
        
        if existing_access:
            return jsonify({
                'success': False,
                'message': f'User {user.username} already has access to {company.name}'
            }), 409
        
        # Create new access record
        new_access = UserCompanyAccess(
            user_id=user_id,
            company_id=company_id,
            created_at=datetime.now(),
            modified_at=datetime.now()
        )
        db.session.add(new_access)
        db.session.commit()
        
        # Log the action
        log_audit(
            action='ADD_USER_COMPANY_ACCESS',
            resource_type='UserCompanyAccess',
            resource_id=str(new_access.id),
            changes=json.dumps({
                'user_id': user_id,
                'user_name': user.username,
                'company_id': str(company_id),
                'company_name': company.name
            })
        )
        
        return jsonify({
            'success': True,
            'message': f'Company {company.name} added to {user.username} successfully',
            'access': {
                'id': str(new_access.id),
                'company_id': str(company_id),
                'company_name': company.name,
                'added_at': new_access.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@access_control_bp.route('/api/remove-company-from-user', methods=['POST'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def remove_company_from_user():
    """API: Remove company access from a user"""
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        company_id = data.get('company_id')
        
        if not user_id or not company_id:
            return jsonify({'success': False, 'message': 'User ID and Company ID are required'}), 400
        
        user = User.query.get(user_id)
        company = Company.query.get(company_id)
        
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        if not company:
            return jsonify({'success': False, 'message': 'Company not found'}), 404
        
        # Verify tenant access - consistent for all roles
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        user_tenant_id = user.organization.tenant_id if user.organization else None
        company_tenant_id = company.tenant_id
        
        if current_tenant_id != user_tenant_id or current_tenant_id != company_tenant_id:
            return jsonify({'success': False, 'message': 'Unauthorized: User and Company must be in your tenant'}), 403
        
        # Find the access record
        access = UserCompanyAccess.query.filter_by(
            user_id=user_id,
            company_id=company_id
        ).first()
        
        if not access:
            return jsonify({
                'success': False,
                'message': f'User {user.username} does not have access to {company.name}'
            }), 404
        
        access_id = access.id
        db.session.delete(access)
        db.session.commit()
        
        # Log the action
        log_audit(
            action='REMOVE_USER_COMPANY_ACCESS',
            resource_type='UserCompanyAccess',
            resource_id=str(access_id),
            changes=json.dumps({
                'user_id': user_id,
                'user_name': user.username,
                'company_id': str(company_id),
                'company_name': company.name
            })
        )
        
        return jsonify({
            'success': True,
            'message': f'Company {company.name} removed from {user.username} successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@access_control_bp.route('/api/get-available-companies/<int:user_id>', methods=['GET'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def get_available_companies(user_id):
    """API: Get companies NOT assigned to the user (for dropdown)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get current user's tenant - always filter by tenant for consistency
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        user_tenant_id = user.organization.tenant_id if user.organization else None
        
        # Verify user belongs to same tenant as current user
        if current_tenant_id != user_tenant_id:
            return jsonify({'success': False, 'message': 'Unauthorized: User is not in your tenant'}), 403
        
        # Get user's current companies
        user_company_ids = [access.company_id for access in user.company_access]
        
        # Get companies from same tenant only (not assigned to user)
        available_companies = Company.query.filter(
            Company.tenant_id == current_tenant_id,
            ~Company.id.in_(user_company_ids)
        ).order_by(Company.name).all()
        
        companies = [
            {
                'id': str(company.id),
                'name': company.name
            }
            for company in available_companies
        ]
        
        return jsonify({
            'success': True,
            'companies': companies,
            'total_available': len(companies)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# =====================================================================
# ROUTE: Toggle User Active Status
# =====================================================================

@access_control_bp.route('/api/toggle-user-status/<int:user_id>', methods=['POST'])
@login_required
@require_role(['Super Admin', 'Tenant Admin', 'HR Manager'])
def toggle_user_status(user_id):
    """API: Toggle user active/inactive status"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        # Get current user's tenant - verify authorization
        current_tenant_id = current_user.organization.tenant_id if current_user.organization else None
        user_tenant_id = user.organization.tenant_id if user.organization else None
        
        # Only Super Admin can manage users from other tenants; Tenant Admin/HR Manager only within their tenant
        if current_user.role.name != 'Super Admin' and current_tenant_id != user_tenant_id:
            return jsonify({'success': False, 'message': 'Unauthorized: User is not in your tenant'}), 403
        
        # Don't allow deactivating self
        if user_id == current_user.id and not user.is_active:
            return jsonify({'success': False, 'message': 'Cannot deactivate your own account'}), 400
        
        # Toggle status
        old_status = user.is_active
        user.is_active = not user.is_active
        db.session.commit()
        
        # Log the action
        log_audit(
            action='TOGGLE_USER_STATUS',
            resource_type='User',
            resource_id=str(user_id),
            changes=json.dumps({
                'user_id': user_id,
                'user_name': user.username,
                'old_status': old_status,
                'new_status': user.is_active
            })
        )
        
        return jsonify({
            'success': True,
            'message': f'User {user.username} is now {"Active" if user.is_active else "Inactive"}',
            'is_active': user.is_active
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# Register blueprint at module load time (before first request)
app.register_blueprint(access_control_bp)


# Initialize when routes are loaded (within app context)
try:
    with app.app_context():
        initialize_access_control_matrix()
except RuntimeError:
    # App context not available during import, will initialize on first request
    pass
