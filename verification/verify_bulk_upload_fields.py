import os
import sys

# Add root directory to path to ensure imports work correctly
sys.path.append(os.getcwd())

import io
import openpyxl
from app import app, db

# Enable Debugging
app.config['DEBUG'] = True
app.config['TESTING'] = True
app.config['PROPAGATE_EXCEPTIONS'] = True
app.config['SESSION_COOKIE_SECURE'] = False # Disable secure cookies for test client

@app.errorhandler(500)
def internal_error(error):
    import traceback
    traceback.print_exc()
    return "Internal Server Error", 500

import routes.routes # Register login/dashboard
import routes.routes_approvals # Register approval dashboard
import routes.routes_tenant_company # Register tenant/company routes
import routes.routes_bulk_upload # Register the routes!
import routes.routes_hr_manager # Register HR dashboard
import routes.routes_masters
import routes.routes_leave
import routes.routes_employee_group
import routes.designation_routes
import routes.routes_ot
import routes.routes_payroll_config
# Add any others needed for UI rendering
from routes.routes_monitoring import monitoring_bp
app.register_blueprint(monitoring_bp)
from core.models import User, Role, Employee, Company, Designation, EmployeeGroup
from flask_login import login_user, current_user

@app.route('/debug/me')
def debug_me():
    return f"User: {current_user}, Auth: {current_user.is_authenticated}"

def verify_bulk_upload():
    print("--- Verifying Bulk Upload & Export Fields ---")
    
    with app.app_context():
        # Setup: Get a Super Admin user
        admin_role = Role.query.filter_by(name='Super Admin').first()
        if not admin_role:
             # Fallback to Tenant Admin or HR Manager if Super Admin not seeded
             admin_role = Role.query.filter(Role.name.in_(['Tenant Admin', 'HR Manager'])).first()
        
        user = User.query.filter_by(role_id=admin_role.id).first()
        if not user:
            print("❌ No Admin user found. Cannot proceed with test.")
            return

        print(f"Using User: {user.username} (Role: {user.role.name})")

        # Setup: Ensure Company and Master Data Exists
        company = Company.query.first()
        if not company:
            print("No Company found.")
            return

        # Ensure Designation and Employee Group exist for testing
        designation = Designation.query.filter_by(name='Test Engineer').first()
        if not designation:
            designation = Designation(name='Test Engineer', description='For Testing')
            db.session.add(designation)
            db.session.commit()
            print("Created Test Designation")

        emp_group = EmployeeGroup.query.filter_by(name='Test Group', company_id=company.id).first()
        if not emp_group:
            emp_group = EmployeeGroup(name='Test Group', category='Test', company_id=company.id)
            db.session.add(emp_group)
            db.session.commit()
            print("Created Test Employee Group")

        client = app.test_client()

        # Login
        with client.session_transaction() as sess:
            sess['_user_id'] = str(user.id)
            sess['_fresh'] = True

        print("Checking Auth Status:")
        print(client.get('/debug/me').get_data(as_text=True))

        print("Refreshed URL Map:")
        print(app.url_map)
        
        with app.test_request_context():
            from flask import url_for
            try:
                print(f"URL for export_employees: {url_for('export_employees')}")
            except Exception as e:
                print(f"Failed to resolve export_employees: {e}")

        # --- Test 1: Downloading Template ---
        print("\nTest 1: Downloading Template...")
        response = client.get('/employees/bulk-upload/download-template')
        
        if response.status_code != 200:
            print(f"Failed to download template. Status: {response.status_code}")
            print(f"Response: {response.get_data(as_text=True)}")
            return
        
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        
        # Check Headers (Row 1)
        headers = [cell.value for cell in ws[1]]
        required_new_headers = [
            'Father Name', 'Location', 'Timezone', 'Designation', 
            'Work Permit Number', 'Hazmat Expiry', 'Airport Pass Expiry',
            'PSA Pass Number', 'PSA Pass Expiry', 'Swift Code', 'IFSC Code',
            'Overtime Group', 'Employee Group'
        ]
        
        missing_headers = [h for h in required_new_headers if h not in headers]
        if missing_headers:
            print(f"Template Headers Missing: {missing_headers}")
        else:
            print("Template Headers Validated (All new fields present)")


        # --- Test 2: Upload Data with New Fields ---
        print("\nTest 2: Uploading Data with New Fields...")
        
        # Construct Excel File
        wb_upload = openpyxl.Workbook()
        ws_upload = wb_upload.active
        ws_upload.title = "Employee Template"
        
        # Write Headers (copy from template)
        for col_num, header in enumerate(headers, 1):
            ws_upload.cell(row=1, column=col_num, value=header)
            
        # Write Data Row
        # Need to map headers to values correctly
        row_data = {
            'Employee ID*': 'TEST_BULK_001',
            'First Name*': 'Test',
            'Last Name*': 'BulkUpload',
            'Father Name': 'Mr. Test Father',
            'Email*': 'test.bulk@example.com',
            'NRIC/Passport*': 'T1234567A',
            'Date of Birth': '1990-01-01',
            'Company Code*': company.code,
            'User Role*': 'User',
            'Hire Date*': '2024-01-01',
            'Employment Type*': 'Full-time',
            'Work Permit Type*': 'Work Permit',
            'Work Permit Number': 'WP123456',
            'Work Permit Expiry': '2025-01-01',
            'Hazmat Expiry': '2025-06-01',
            'Airport Pass Expiry': '2025-07-01',
            'PSA Pass Number': 'PSA999',
            'PSA Pass Expiry': '2025-08-01',
            'Basic Salary*': '3000',
            'Designation': 'Test Engineer',
            'Employee Group': 'Test Group',
            'Location': 'HQ',
            'Timezone': 'Asia/Singapore',
            'Swift Code': 'SWIFTXX',
            'IFSC Code': 'IFSCXX',
            'Bank Name': 'DBS',
            'Bank Account': '123-456-789',
            'Account Holder Name': 'Test Bulk',
            'Overtime Group': 'Group A'
        }
        
        for col_num, header in enumerate(headers, 1):
            val = row_data.get(header, '')
            if header == 'Employer CPF Rate' or header == 'Employee CPF Rate':
                val = '0' # Defaults
            ws_upload.cell(row=3, column=col_num, value=val)

        # Save to BytesIO
        file_io = io.BytesIO()
        wb_upload.save(file_io)
        file_io.seek(0)
        
        data = {
            'excel_file': (file_io, 'test_upload.xlsx')
        }
        
        response = client.post('/employees/bulk-upload/upload', data=data, content_type='multipart/form-data', follow_redirects=False)
        
        # Check for redirect (302) which implies success or handled error
        if response.status_code == 302:
             print("Upload API returned 302 Redirect (Likely Success)")
        else:
             print(f"Warning: Upload finished with status {response.status_code}. Response: {response.get_data(as_text=True)[:200]}...")

        # Verify DB
        emp = Employee.query.filter_by(email='test.bulk@example.com').first()
        if not emp:
            print("Employee not created in DB")
        else:
            errors = []
            if emp.father_name != 'Mr. Test Father': errors.append(f"Father Name: {emp.father_name}")
            if emp.work_permit_number != 'WP123456': errors.append(f"Work Permit Number: {emp.work_permit_number}")
            if emp.designation is None or emp.designation.name != 'Test Engineer': errors.append(f"Designation: {emp.designation}")
            if emp.employee_group is None or emp.employee_group.name != 'Test Group': errors.append(f"Employee Group: {emp.employee_group}")
            if str(emp.hazmat_expiry) != '2025-06-01': errors.append(f"Hazmat Expiry: {emp.hazmat_expiry}")
            if emp.swift_code != 'SWIFTXX': errors.append(f"Swift Code: {emp.swift_code}")
            
            if errors:
                print(f"Data Verification Failed. Mismatches: {errors}")
            else:
                print("Data Verification Successful (All fields matched)")

        # --- Test 3: Export Data ---
        print("\nTest 3: Exporting Data...")
        response = client.get('/employees/export')
        
        if response.status_code != 200:
            print(f"Export failed. Status: {response.status_code}")
            return
            
        wb_export = openpyxl.load_workbook(io.BytesIO(response.data))
        ws_export = wb_export.active
        
        # Find row for our test employee
        header_map = {cell.value: i for i, cell in enumerate(ws_export[1], 1)}
        
        found = False
        for row in ws_export.iter_rows(min_row=3, values_only=True):
            # Column 4 in template is Email? Headers list says: ID, First, Last, Father, Email (col 5)
            # Let's use header map to be safe
            email_idx = header_map.get('Email*')
            if email_idx and row[email_idx-1] == 'test.bulk@example.com':
                found = True
                print("Found exported row for test employee")
                
                # Check specific field
                father_idx = header_map.get('Father Name')
                designation_idx = header_map.get('Designation')
                wp_num_idx = header_map.get('Work Permit Number')
                
                if father_idx and row[father_idx-1] == 'Mr. Test Father':
                    print("  - Father Name Exported Correctly")
                else:
                     print(f"  - Father Name Mismatch: {row[father_idx-1] if father_idx else 'N/A'}")

                if designation_idx and row[designation_idx-1] == 'Test Engineer':
                    print("  - Designation Exported Correctly")
                else:
                     print(f"  - Designation Mismatch: {row[designation_idx-1] if designation_idx else 'N/A'}")
                break
        
        if not found:
            print("Exported data does not contain the test employee")

if __name__ == "__main__":
    verify_bulk_upload()
